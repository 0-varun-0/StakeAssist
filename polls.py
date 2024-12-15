from flask import Flask, request, jsonify
from openpyxl import load_workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'polls.xlsx'

# Ensure the Excel file exists
if not os.path.exists(EXCEL_FILE):
    # Create a new workbook and sheet if file doesn't exist
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Polls"
    sheet.append(["Question", "Options"])  # Headers
    wb.save(EXCEL_FILE)

@app.route('/add-poll', methods=['POST'])
def add_poll():
    """Add a new poll to the Excel file."""
    data = request.get_json()
    question = data.get("question")
    options = data.get("options")

    if not question or not options or len(options) < 2:
        return jsonify({"error": "Invalid input. Provide a question and at least two options."}), 400

    wb = load_workbook(EXCEL_FILE)
    sheet = wb["Polls"]

    # Append the new poll
    sheet.append([question, ",".join(options)])
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Poll added successfully!"})

if __name__ == '__main__':
    app.run(port=3000, debug=True)
